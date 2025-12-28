from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from django.utils import timezone
from django.db.models import Q, Sum, Case, When, DecimalField, Value, F
from django.db.models.functions import TruncMonth, Coalesce
from django.http import HttpResponse
from drf_spectacular.utils import extend_schema
from datetime import date, datetime
from dateutil.relativedelta import relativedelta
from decimal import Decimal
import csv
import io
from .serializers import (
    MessageSerializer,
    MemberSerializer,
    MemberRegistrationSerializer,
    MemberLoginSerializer,
    TransactionSerializer,
    CategorySerializer,
    UserSettingsSerializer,
    DashboardStatsSerializer,
    DashboardDynamicsSerializer,
    TopCategoriesSerializer,
    ProfitLossReportSerializer,
    CashFlowReportSerializer,
    TaxReportSerializer
)
from .models import Member, Transaction, Category, UserSettings
from .authentication import SessionAuthentication
from .permissions import IsAuthenticated


class HelloView(APIView):
    """
    A simple API endpoint that returns a greeting message.
    """

    @extend_schema(
        responses={200: MessageSerializer}, description="Get a hello world message"
    )
    def get(self, request):
        data = {"message": "Hello!", "timestamp": timezone.now()}
        serializer = MessageSerializer(data)
        return Response(serializer.data)


class RegisterView(APIView):
    """
    Register new member and automatically login
    """
    authentication_classes = []
    permission_classes = []

    @extend_schema(
        request=MemberRegistrationSerializer,
        responses={
            201: MemberSerializer,
            400: {'type': 'object', 'properties': {'error': {'type': 'string'}}},
            409: {'type': 'object', 'properties': {'error': {'type': 'string'}}}
        },
        description="Register new member and automatically login"
    )
    def post(self, request):
        serializer = MemberRegistrationSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response({'error': str(serializer.errors)}, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if email already exists
        if Member.objects.filter(email=serializer.validated_data['email']).exists():
            return Response(
                {'error': 'Member with this email already exists'},
                status=status.HTTP_409_CONFLICT
            )
        
        member = serializer.save()
        
        # Automatically login after registration
        request.session['member_id'] = member.id
        
        response_serializer = MemberSerializer(member)
        response = Response(response_serializer.data, status=status.HTTP_201_CREATED)
        
        # Set HttpOnly cookie
        response.set_cookie(
            key='sessionid',
            value=request.session.session_key,
            httponly=True,
            samesite='Lax',
            path='/'
        )
        
        return response


class LoginView(APIView):
    """
    Login member and set session cookie
    """
    authentication_classes = []
    permission_classes = []

    @extend_schema(
        request=MemberLoginSerializer,
        responses={
            200: MemberSerializer,
            400: {'type': 'object', 'properties': {'error': {'type': 'string'}}},
            401: {'type': 'object', 'properties': {'error': {'type': 'string'}}}
        },
        description="Authenticate member and set session cookie"
    )
    def post(self, request):
        serializer = MemberLoginSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response({'error': str(serializer.errors)}, status=status.HTTP_400_BAD_REQUEST)
        
        email = serializer.validated_data['email']
        password = serializer.validated_data['password']
        
        try:
            member = Member.objects.get(email=email)
        except Member.DoesNotExist:
            return Response(
                {'error': 'Invalid credentials'},
                status=status.HTTP_401_UNAUTHORIZED
            )
        
        if not member.check_password(password):
            return Response(
                {'error': 'Invalid credentials'},
                status=status.HTTP_401_UNAUTHORIZED
            )
        
        # Create session
        request.session['member_id'] = member.id
        
        response_serializer = MemberSerializer(member)
        response = Response(response_serializer.data, status=status.HTTP_200_OK)
        
        # Set HttpOnly cookie
        response.set_cookie(
            key='sessionid',
            value=request.session.session_key,
            httponly=True,
            samesite='Lax',
            path='/'
        )
        
        return response


class LogoutView(APIView):
    """
    Logout current member and clear session cookie
    """
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    @extend_schema(
        responses={
            200: {'type': 'object', 'properties': {'message': {'type': 'string'}}},
            401: {'type': 'object', 'properties': {'error': {'type': 'string'}}}
        },
        description="Logout current member and clear session cookie"
    )
    def post(self, request):
        # Clear session
        request.session.flush()
        
        response = Response({'message': 'Successfully logged out'}, status=status.HTTP_200_OK)
        
        # Clear cookie
        response.delete_cookie('sessionid', path='/')
        
        return response


class MeView(APIView):
    """
    Get information about currently authenticated member
    """
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    @extend_schema(
        responses={
            200: MemberSerializer,
            401: {'type': 'object', 'properties': {'error': {'type': 'string'}}}
        },
        description="Get information about currently authenticated member"
    )
    def get(self, request):
        serializer = MemberSerializer(request.user)
        return Response(serializer.data, status=status.HTTP_200_OK)


class TransactionPagination(PageNumberPagination):
    """
    Pagination for transactions
    """
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 1000


class TransactionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing transactions
    """
    serializer_class = TransactionSerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]
    pagination_class = TransactionPagination

    def get_queryset(self):
        """
        Get transactions for current user with filtering
        """
        queryset = Transaction.objects.filter(member=self.request.user)
        
        # Filter by type
        transaction_type = self.request.query_params.get('type', None)
        if transaction_type:
            queryset = queryset.filter(type=transaction_type)
        
        # Filter by category
        category_id = self.request.query_params.get('category_id', None)
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        
        # Filter by date range
        date_from = self.request.query_params.get('date_from', None)
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        
        date_to = self.request.query_params.get('date_to', None)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        
        # Filter by counterparty
        counterparty = self.request.query_params.get('counterparty', None)
        if counterparty:
            queryset = queryset.filter(counterparty__icontains=counterparty)
        
        # Filter by project
        project = self.request.query_params.get('project', None)
        if project:
            queryset = queryset.filter(project__icontains=project)
        
        # Filter by account
        account = self.request.query_params.get('account', None)
        if account:
            queryset = queryset.filter(account__icontains=account)
        
        # Search in description, counterparty, project
        search = self.request.query_params.get('search', None)
        if search:
            queryset = queryset.filter(
                Q(description__icontains=search) |
                Q(counterparty__icontains=search) |
                Q(project__icontains=search)
            )
        
        return queryset.select_related('category')

    def perform_create(self, serializer):
        """
        Set member to current user on create
        """
        serializer.save(member=self.request.user)

    def perform_update(self, serializer):
        """
        Ensure member stays the same on update
        """
        serializer.save(member=self.request.user)

    @action(detail=False, methods=['get'], url_path='export')
    def export(self, request):
        """
        Export transactions to CSV or Excel format
        """
        export_format = request.query_params.get('format', None)
        
        if not export_format:
            return Response(
                {'error': 'Format parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if export_format not in ['csv', 'excel']:
            return Response(
                {'error': 'Invalid format. Allowed: csv, excel'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get filtered queryset
        queryset = self.get_queryset()
        
        if export_format == 'csv':
            return self._export_csv(queryset)
        else:
            return self._export_excel(queryset)

    def _export_csv(self, queryset):
        """
        Export transactions to CSV
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            'ID', 'Type', 'Amount', 'Date', 'Category', 'Description',
            'Counterparty', 'Project', 'Account', 'Document', 'Created At'
        ])
        
        # Write data
        for transaction in queryset:
            writer.writerow([
                transaction.id,
                transaction.type,
                str(transaction.amount),
                transaction.date.strftime('%Y-%m-%d'),
                transaction.category.name,
                transaction.description or '',
                transaction.counterparty or '',
                transaction.project or '',
                transaction.account or '',
                transaction.document or '',
                transaction.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        response = HttpResponse(output.getvalue(), content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="transactions.csv"'
        return response

    def _export_excel(self, queryset):
        """
        Export transactions to Excel format
        """
        try:
            import openpyxl
            from openpyxl.styles import Font
        except ImportError:
            return Response(
                {'error': 'Excel export is not available'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Transactions'
        
        # Write header
        headers = [
            'ID', 'Type', 'Amount', 'Date', 'Category', 'Description',
            'Counterparty', 'Project', 'Account', 'Document', 'Created At'
        ]
        worksheet.append(headers)
        
        # Make header bold
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        
        # Write data
        for transaction in queryset:
            worksheet.append([
                transaction.id,
                transaction.type,
                float(transaction.amount),
                transaction.date.strftime('%Y-%m-%d'),
                transaction.category.name,
                transaction.description or '',
                transaction.counterparty or '',
                transaction.project or '',
                transaction.account or '',
                transaction.document or '',
                transaction.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
        return response


class CategoryViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing categories
    """
    serializer_class = CategorySerializer
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        """
        Get categories for current user with filtering
        """
        queryset = Category.objects.filter(member=self.request.user)
        
        # Filter by type
        category_type = self.request.query_params.get('type', None)
        if category_type:
            queryset = queryset.filter(type=category_type)
        
        return queryset

    def perform_create(self, serializer):
        """
        Set member to current user on create
        """
        serializer.save(member=self.request.user)

    def perform_update(self, serializer):
        """
        Ensure member stays the same on update
        """
        serializer.save(member=self.request.user)

    def destroy(self, request, *args, **kwargs):
        """
        Prevent deletion of default categories
        """
        category = self.get_object()
        
        if category.is_default:
            return Response(
                {'error': 'Cannot delete default category'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return super().destroy(request, *args, **kwargs)


class UserSettingsViewSet(viewsets.ViewSet):
    """
    ViewSet for managing user settings
    """
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    def retrieve(self, request):
        """
        Get current user settings
        """
        try:
            settings = UserSettings.objects.get(member=request.user)
        except UserSettings.DoesNotExist:
            # Create default settings if they don't exist
            settings = UserSettings.objects.create(member=request.user)
        
        serializer = UserSettingsSerializer(settings)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def update(self, request):
        """
        Update current user settings
        """
        try:
            settings = UserSettings.objects.get(member=request.user)
        except UserSettings.DoesNotExist:
            # Create default settings if they don't exist
            settings = UserSettings.objects.create(member=request.user)
        
        serializer = UserSettingsSerializer(settings, data=request.data, partial=True)
        
        if not serializer.is_valid():
            return Response(
                {'error': str(serializer.errors)},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)


class PeriodMixin:
    """
    Mixin to handle period calculation
    """
    def get_date_range(self, request):
        """
        Calculate date range based on period parameter
        """
        period = request.query_params.get('period', 'current_month')
        date_from_str = request.query_params.get('date_from')
        date_to_str = request.query_params.get('date_to')
        
        today = date.today()
        
        if period == 'current_month':
            date_from = today.replace(day=1)
            next_month = date_from + relativedelta(months=1)
            date_to = next_month - relativedelta(days=1)
        elif period == 'last_month':
            first_day_current = today.replace(day=1)
            date_to = first_day_current - relativedelta(days=1)
            date_from = date_to.replace(day=1)
        elif period == 'current_year':
            date_from = today.replace(month=1, day=1)
            date_to = today.replace(month=12, day=31)
        elif period == 'custom':
            if not date_from_str or not date_to_str:
                raise ValueError('date_from and date_to are required for custom period')
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                raise ValueError('Invalid date format. Use YYYY-MM-DD')
        else:
            raise ValueError('Invalid period parameter')
        
        return period, date_from, date_to


class DashboardStatsView(APIView, PeriodMixin):
    """
    Get dashboard statistics
    """
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    @extend_schema(
        responses={
            200: DashboardStatsSerializer,
            400: {'type': 'object', 'properties': {'error': {'type': 'string'}}},
            401: {'type': 'object', 'properties': {'error': {'type': 'string'}}}
        },
        description="Get dashboard statistics"
    )
    def get(self, request):
        try:
            period, date_from, date_to = self.get_date_range(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get user settings for tax rate
        try:
            settings = UserSettings.objects.get(member=request.user)
            tax_rate = settings.tax_rate
        except UserSettings.DoesNotExist:
            tax_rate = Decimal('0.00')
        
        # Calculate totals
        transactions = Transaction.objects.filter(
            member=request.user,
            date__gte=date_from,
            date__lte=date_to
        )
        
        total_income = transactions.filter(type='income').aggregate(
            total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        )['total']
        
        total_expenses = transactions.filter(type='expense').aggregate(
            total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        )['total']
        
        # Calculate taxes
        taxable_income = total_income - total_expenses
        taxes = (taxable_income * tax_rate / 100) if taxable_income > 0 else Decimal('0.00')
        
        # Calculate metrics
        cash_flow = total_income - total_expenses
        net_profit = cash_flow - taxes
        profitability = (net_profit / total_income * 100) if total_income > 0 else Decimal('0.00')
        
        data = {
            'total_income': total_income,
            'total_expenses': total_expenses,
            'net_profit': net_profit,
            'taxes': taxes,
            'cash_flow': cash_flow,
            'profitability': profitability,
            'period': period,
            'date_from': date_from,
            'date_to': date_to
        }
        
        serializer = DashboardStatsSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)


class DashboardDynamicsView(APIView, PeriodMixin):
    """
    Get dynamics by months
    """
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    @extend_schema(
        responses={
            200: DashboardDynamicsSerializer,
            400: {'type': 'object', 'properties': {'error': {'type': 'string'}}},
            401: {'type': 'object', 'properties': {'error': {'type': 'string'}}}
        },
        description="Get dynamics by months"
    )
    def get(self, request):
        try:
            period, date_from, date_to = self.get_date_range(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get monthly aggregated data
        transactions = Transaction.objects.filter(
            member=request.user,
            date__gte=date_from,
            date__lte=date_to
        ).annotate(
            month=TruncMonth('date')
        ).values('month').annotate(
            income=Coalesce(
                Sum(Case(
                    When(type='income', then='amount'),
                    default=Value(0),
                    output_field=DecimalField()
                )),
                Value(0),
                output_field=DecimalField()
            ),
            expenses=Coalesce(
                Sum(Case(
                    When(type='expense', then='amount'),
                    default=Value(0),
                    output_field=DecimalField()
                )),
                Value(0),
                output_field=DecimalField()
            )
        ).order_by('month')
        
        # Calculate profit for each month
        dynamics = []
        for item in transactions:
            profit = item['income'] - item['expenses']
            dynamics.append({
                'month': item['month'],
                'income': item['income'],
                'expenses': item['expenses'],
                'profit': profit
            })
        
        data = {
            'dynamics': dynamics,
            'period': period,
            'date_from': date_from,
            'date_to': date_to
        }
        
        serializer = DashboardDynamicsSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)


class TopCategoriesView(APIView, PeriodMixin):
    """
    Get top categories
    """
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    @extend_schema(
        responses={
            200: TopCategoriesSerializer,
            400: {'type': 'object', 'properties': {'error': {'type': 'string'}}},
            401: {'type': 'object', 'properties': {'error': {'type': 'string'}}}
        },
        description="Get top categories"
    )
    def get(self, request):
        try:
            period, date_from, date_to = self.get_date_range(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        limit = int(request.query_params.get('limit', 5))
        
        # Get top income categories
        income_categories = Transaction.objects.filter(
            member=request.user,
            type='income',
            date__gte=date_from,
            date__lte=date_to
        ).values(
            'category_id',
            'category__name'
        ).annotate(
            amount=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        ).order_by('-amount')[:limit]
        
        # Calculate total income for percentages
        total_income = sum(item['amount'] for item in income_categories)
        
        income_data = []
        for item in income_categories:
            percentage = (item['amount'] / total_income * 100) if total_income > 0 else Decimal('0.00')
            income_data.append({
                'category_id': item['category_id'],
                'category_name': item['category__name'],
                'amount': item['amount'],
                'percentage': percentage
            })
        
        # Get top expense categories
        expense_categories = Transaction.objects.filter(
            member=request.user,
            type='expense',
            date__gte=date_from,
            date__lte=date_to
        ).values(
            'category_id',
            'category__name'
        ).annotate(
            amount=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        ).order_by('-amount')[:limit]
        
        # Calculate total expenses for percentages
        total_expenses = sum(item['amount'] for item in expense_categories)
        
        expense_data = []
        for item in expense_categories:
            percentage = (item['amount'] / total_expenses * 100) if total_expenses > 0 else Decimal('0.00')
            expense_data.append({
                'category_id': item['category_id'],
                'category_name': item['category__name'],
                'amount': item['amount'],
                'percentage': percentage
            })
        
        data = {
            'income_categories': income_data,
            'expense_categories': expense_data,
            'period': period,
            'date_from': date_from,
            'date_to': date_to
        }
        
        serializer = TopCategoriesSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ProfitLossReportView(APIView, PeriodMixin):
    """
    Get profit and loss report
    """
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    @extend_schema(
        responses={
            200: ProfitLossReportSerializer,
            400: {'type': 'object', 'properties': {'error': {'type': 'string'}}},
            401: {'type': 'object', 'properties': {'error': {'type': 'string'}}}
        },
        description="Get profit and loss report"
    )
    def get(self, request):
        try:
            period, date_from, date_to = self.get_date_range(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get user settings for tax rate
        try:
            settings = UserSettings.objects.get(member=request.user)
            tax_rate = settings.tax_rate
        except UserSettings.DoesNotExist:
            tax_rate = Decimal('0.00')
        
        # Get income by category
        income_categories = Transaction.objects.filter(
            member=request.user,
            type='income',
            date__gte=date_from,
            date__lte=date_to
        ).values(
            'category_id',
            'category__name'
        ).annotate(
            amount=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        ).order_by('-amount')
        
        total_income = sum(item['amount'] for item in income_categories)
        
        income_data = {
            'total': total_income,
            'categories': [
                {
                    'category_id': item['category_id'],
                    'category_name': item['category__name'],
                    'amount': item['amount']
                }
                for item in income_categories
            ]
        }
        
        # Get expenses by category
        expense_categories = Transaction.objects.filter(
            member=request.user,
            type='expense',
            date__gte=date_from,
            date__lte=date_to
        ).values(
            'category_id',
            'category__name'
        ).annotate(
            amount=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        ).order_by('-amount')
        
        total_expenses = sum(item['amount'] for item in expense_categories)
        
        expenses_data = {
            'total': total_expenses,
            'categories': [
                {
                    'category_id': item['category_id'],
                    'category_name': item['category__name'],
                    'amount': item['amount']
                }
                for item in expense_categories
            ]
        }
        
        # Calculate metrics
        gross_profit = total_income - total_expenses
        taxable_income = gross_profit
        taxes = (taxable_income * tax_rate / 100) if taxable_income > 0 else Decimal('0.00')
        net_profit = gross_profit - taxes
        
        data = {
            'income': income_data,
            'expenses': expenses_data,
            'gross_profit': gross_profit,
            'taxes': taxes,
            'net_profit': net_profit,
            'period': period,
            'date_from': date_from,
            'date_to': date_to
        }
        
        serializer = ProfitLossReportSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)


class CashFlowReportView(APIView, PeriodMixin):
    """
    Get cash flow report
    """
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    @extend_schema(
        responses={
            200: CashFlowReportSerializer,
            400: {'type': 'object', 'properties': {'error': {'type': 'string'}}},
            401: {'type': 'object', 'properties': {'error': {'type': 'string'}}}
        },
        description="Get cash flow report"
    )
    def get(self, request):
        try:
            period, date_from, date_to = self.get_date_range(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get total cash flows
        transactions = Transaction.objects.filter(
            member=request.user,
            date__gte=date_from,
            date__lte=date_to
        )
        
        cash_inflows = transactions.filter(type='income').aggregate(
            total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        )['total']
        
        cash_outflows = transactions.filter(type='expense').aggregate(
            total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        )['total']
        
        net_cash_flow = cash_inflows - cash_outflows
        
        # Get monthly data
        monthly_transactions = transactions.annotate(
            month=TruncMonth('date')
        ).values('month').annotate(
            inflows=Coalesce(
                Sum(Case(
                    When(type='income', then='amount'),
                    default=Value(0),
                    output_field=DecimalField()
                )),
                Value(0),
                output_field=DecimalField()
            ),
            outflows=Coalesce(
                Sum(Case(
                    When(type='expense', then='amount'),
                    default=Value(0),
                    output_field=DecimalField()
                )),
                Value(0),
                output_field=DecimalField()
            )
        ).order_by('month')
        
        monthly_data = []
        for item in monthly_transactions:
            net_flow = item['inflows'] - item['outflows']
            monthly_data.append({
                'month': item['month'],
                'inflows': item['inflows'],
                'outflows': item['outflows'],
                'net_flow': net_flow
            })
        
        data = {
            'cash_inflows': cash_inflows,
            'cash_outflows': cash_outflows,
            'net_cash_flow': net_cash_flow,
            'monthly_data': monthly_data,
            'period': period,
            'date_from': date_from,
            'date_to': date_to
        }
        
        serializer = CashFlowReportSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)


class TaxReportView(APIView, PeriodMixin):
    """
    Get tax report
    """
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    @extend_schema(
        responses={
            200: TaxReportSerializer,
            400: {'type': 'object', 'properties': {'error': {'type': 'string'}}},
            401: {'type': 'object', 'properties': {'error': {'type': 'string'}}}
        },
        description="Get tax report"
    )
    def get(self, request):
        try:
            period, date_from, date_to = self.get_date_range(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        # Get user settings for tax rate
        try:
            settings = UserSettings.objects.get(member=request.user)
            tax_rate = settings.tax_rate
        except UserSettings.DoesNotExist:
            tax_rate = Decimal('0.00')
        
        # Get totals
        transactions = Transaction.objects.filter(
            member=request.user,
            date__gte=date_from,
            date__lte=date_to
        )
        
        total_income = transactions.filter(type='income').aggregate(
            total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        )['total']
        
        total_expenses = transactions.filter(type='expense').aggregate(
            total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        )['total']
        
        taxable_income = total_income - total_expenses
        estimated_tax = (taxable_income * tax_rate / 100) if taxable_income > 0 else Decimal('0.00')
        
        # Get monthly breakdown
        monthly_transactions = transactions.annotate(
            month=TruncMonth('date')
        ).values('month').annotate(
            income=Coalesce(
                Sum(Case(
                    When(type='income', then='amount'),
                    default=Value(0),
                    output_field=DecimalField()
                )),
                Value(0),
                output_field=DecimalField()
            ),
            expenses=Coalesce(
                Sum(Case(
                    When(type='expense', then='amount'),
                    default=Value(0),
                    output_field=DecimalField()
                )),
                Value(0),
                output_field=DecimalField()
            )
        ).order_by('month')
        
        monthly_breakdown = []
        for item in monthly_transactions:
            monthly_taxable_income = item['income'] - item['expenses']
            monthly_tax = (monthly_taxable_income * tax_rate / 100) if monthly_taxable_income > 0 else Decimal('0.00')
            monthly_breakdown.append({
                'month': item['month'],
                'income': item['income'],
                'expenses': item['expenses'],
                'taxable_income': monthly_taxable_income,
                'tax': monthly_tax
            })
        
        data = {
            'taxable_income': taxable_income,
            'tax_rate': tax_rate,
            'estimated_tax': estimated_tax,
            'total_income': total_income,
            'total_expenses': total_expenses,
            'monthly_breakdown': monthly_breakdown,
            'period': period,
            'date_from': date_from,
            'date_to': date_to
        }
        
        serializer = TaxReportSerializer(data)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ReportExportView(APIView, PeriodMixin):
    """
    Export report in PDF or Excel format
    """
    authentication_classes = [SessionAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        report_type = request.query_params.get('report_type')
        export_format = request.query_params.get('format')
        
        if not report_type or not export_format:
            return Response(
                {'error': 'report_type and format parameters are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if report_type not in ['profit_loss', 'cash_flow', 'tax']:
            return Response(
                {'error': 'Invalid report_type. Allowed: profit_loss, cash_flow, tax'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if export_format not in ['pdf', 'excel']:
            return Response(
                {'error': 'Invalid format. Allowed: pdf, excel'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            period, date_from, date_to = self.get_date_range(request)
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
        if export_format == 'excel':
            return self._export_excel(report_type, period, date_from, date_to)
        else:
            return self._export_pdf(report_type, period, date_from, date_to)
    
    def _export_excel(self, report_type, period, date_from, date_to):
        """
        Export report to Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
        except ImportError:
            return Response(
                {'error': 'Excel export is not available'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        
        if report_type == 'profit_loss':
            worksheet.title = 'Profit & Loss'
            self._write_profit_loss_excel(worksheet, date_from, date_to)
        elif report_type == 'cash_flow':
            worksheet.title = 'Cash Flow'
            self._write_cash_flow_excel(worksheet, date_from, date_to)
        elif report_type == 'tax':
            worksheet.title = 'Tax Report'
            self._write_tax_excel(worksheet, date_from, date_to)
        
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'
        return response
    
    def _write_profit_loss_excel(self, worksheet, date_from, date_to):
        """
        Write profit and loss data to Excel worksheet
        """
        from openpyxl.styles import Font
        
        worksheet.append(['Profit & Loss Report'])
        worksheet.append([f'Period: {date_from} to {date_to}'])
        worksheet.append([])
        
        # Get data
        try:
            settings = UserSettings.objects.get(member=self.request.user)
            tax_rate = settings.tax_rate
        except UserSettings.DoesNotExist:
            tax_rate = Decimal('0.00')
        
        income_categories = Transaction.objects.filter(
            member=self.request.user,
            type='income',
            date__gte=date_from,
            date__lte=date_to
        ).values('category__name').annotate(
            amount=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        ).order_by('-amount')
        
        expense_categories = Transaction.objects.filter(
            member=self.request.user,
            type='expense',
            date__gte=date_from,
            date__lte=date_to
        ).values('category__name').annotate(
            amount=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        ).order_by('-amount')
        
        total_income = sum(item['amount'] for item in income_categories)
        total_expenses = sum(item['amount'] for item in expense_categories)
        gross_profit = total_income - total_expenses
        taxes = (gross_profit * tax_rate / 100) if gross_profit > 0 else Decimal('0.00')
        net_profit = gross_profit - taxes
        
        # Write income
        worksheet.append(['Income'])
        worksheet['A' + str(worksheet.max_row)].font = Font(bold=True)
        for item in income_categories:
            worksheet.append([item['category__name'], float(item['amount'])])
        worksheet.append(['Total Income', float(total_income)])
        worksheet['A' + str(worksheet.max_row)].font = Font(bold=True)
        worksheet.append([])
        
        # Write expenses
        worksheet.append(['Expenses'])
        worksheet['A' + str(worksheet.max_row)].font = Font(bold=True)
        for item in expense_categories:
            worksheet.append([item['category__name'], float(item['amount'])])
        worksheet.append(['Total Expenses', float(total_expenses)])
        worksheet['A' + str(worksheet.max_row)].font = Font(bold=True)
        worksheet.append([])
        
        # Write summary
        worksheet.append(['Gross Profit', float(gross_profit)])
        worksheet.append(['Taxes', float(taxes)])
        worksheet.append(['Net Profit', float(net_profit)])
        worksheet['A' + str(worksheet.max_row)].font = Font(bold=True)
    
    def _write_cash_flow_excel(self, worksheet, date_from, date_to):
        """
        Write cash flow data to Excel worksheet
        """
        from openpyxl.styles import Font
        
        worksheet.append(['Cash Flow Report'])
        worksheet.append([f'Period: {date_from} to {date_to}'])
        worksheet.append([])
        
        # Get data
        transactions = Transaction.objects.filter(
            member=self.request.user,
            date__gte=date_from,
            date__lte=date_to
        )
        
        cash_inflows = transactions.filter(type='income').aggregate(
            total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        )['total']
        
        cash_outflows = transactions.filter(type='expense').aggregate(
            total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        )['total']
        
        net_cash_flow = cash_inflows - cash_outflows
        
        worksheet.append(['Cash Inflows', float(cash_inflows)])
        worksheet.append(['Cash Outflows', float(cash_outflows)])
        worksheet.append(['Net Cash Flow', float(net_cash_flow)])
        worksheet['A' + str(worksheet.max_row)].font = Font(bold=True)
    
    def _write_tax_excel(self, worksheet, date_from, date_to):
        """
        Write tax data to Excel worksheet
        """
        from openpyxl.styles import Font
        
        worksheet.append(['Tax Report'])
        worksheet.append([f'Period: {date_from} to {date_to}'])
        worksheet.append([])
        
        # Get data
        try:
            settings = UserSettings.objects.get(member=self.request.user)
            tax_rate = settings.tax_rate
        except UserSettings.DoesNotExist:
            tax_rate = Decimal('0.00')
        
        transactions = Transaction.objects.filter(
            member=self.request.user,
            date__gte=date_from,
            date__lte=date_to
        )
        
        total_income = transactions.filter(type='income').aggregate(
            total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        )['total']
        
        total_expenses = transactions.filter(type='expense').aggregate(
            total=Coalesce(Sum('amount'), Value(0), output_field=DecimalField())
        )['total']
        
        taxable_income = total_income - total_expenses
        estimated_tax = (taxable_income * tax_rate / 100) if taxable_income > 0 else Decimal('0.00')
        
        worksheet.append(['Total Income', float(total_income)])
        worksheet.append(['Total Expenses', float(total_expenses)])
        worksheet.append(['Taxable Income', float(taxable_income)])
        worksheet.append(['Tax Rate', f'{float(tax_rate)}%'])
        worksheet.append(['Estimated Tax', float(estimated_tax)])
        worksheet['A' + str(worksheet.max_row)].font = Font(bold=True)
    
    def _export_pdf(self, report_type, period, date_from, date_to):
        """
        Export report to PDF
        """
        return Response(
            {'error': 'PDF export is not implemented yet'},
            status=status.HTTP_400_BAD_REQUEST
        )